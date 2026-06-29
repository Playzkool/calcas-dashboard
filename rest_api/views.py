import csv
import io
import unicodedata
import zipfile
from datetime import date

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import PasswordResetForm
from django.http import HttpResponse
from django.middleware.csrf import get_token
from rest_framework import status
from rest_framework.authentication import SessionAuthentication
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_api.html_export import generate_registration_html

from registration.models import (
    LegalRepresentative,
    PupilLegalRepresentative,
    RegistrationCampaign,
    RegistrationFile,
    RegistrationSupervisor,
)
from rest_api.serializers import (
    LegalRepresentativeCreateSerializer,
    LegalRepresentativeListItemSerializer,
    LegalRepresentativeProfileSerializer,
    RegistrationDetailSerializer,
    RegistrationFileCreateSerializer,
    RegistrationFileUpdateSerializer,
    RegistrationListItemSerializer,
)


EXPORT_HEADERS = [
    "Prénom", "Nom", "Date de naissance", "Lieu de naissance", "Nationalité",
    "Code postal", "Adresse", "Niveau", "Situation familiale", "Frères", "Sœurs",
    "Autorisation SAMU", "Médecin traitant", "Allergies", "Autres vaccins", "Antécédents médicaux",
    "Autorisation sortie pédagogique", "Droit à l'image", "Charte acceptée",
    "Pièce naissance/livret", "Carnet santé/vaccins", "Attestation assurance", "Jugement divorce",
    "Statut dossier",
    "LR1 Email", "LR1 Prénom", "LR1 Nom", "LR1 Tél mobile", "LR1 Tél fixe", "LR1 Autorité parentale",
    "LR2 Email", "LR2 Prénom", "LR2 Nom", "LR2 Tél mobile", "LR2 Tél fixe", "LR2 Autorité parentale",
]

_FAMILY_LABELS = {
    "married_or_cohabiting": "Marié(e) / En couple",
    "divorced_or_separated": "Divorcé(e) / Séparé(e)",
    "single_parent": "Parent seul",
}


def _bool_label(v):
    if v is True:
        return "Oui"
    if v is False:
        return "Non"
    return ""


def _registration_row(reg):
    pupil = reg.pupil
    lrs = [plr.legal_representative for plr in pupil.pupillegalrepresentative_set.all()]
    diseases = [k for k, v in (reg.diseases_history or {}).items() if v]
    row = [
        pupil.firstname or "",
        pupil.lastname or "",
        str(pupil.birth_date) if pupil.birth_date else "",
        pupil.birth_place or "",
        pupil.nationality or "",
        pupil.postal_code or "",
        pupil.address or "",
        pupil.get_grade_display(),
        _FAMILY_LABELS.get(pupil.family_situation, pupil.family_situation or ""),
        str(pupil.siblings_brothers) if pupil.siblings_brothers is not None else "",
        str(pupil.siblings_sisters) if pupil.siblings_sisters is not None else "",
        _bool_label(reg.samu_authorized),
        reg.doctor_name_phone or "",
        reg.allergies_info or "",
        reg.other_vaccines or "",
        ", ".join(diseases),
        _bool_label(reg.school_trips_authorized),
        _bool_label(reg.image_rights_authorized),
        _bool_label(reg.charter_accepted),
        "Oui" if reg.document else "Non",
        "Oui" if reg.vaccination_document else "Non",
        "Oui" if reg.insurance_document else "Non",
        "Oui" if reg.divorce_judgment else "Non",
        "Clôturé" if reg.is_closed else "En cours",
    ]
    for i in range(2):
        if i < len(lrs):
            lr = lrs[i]
            row += [lr.user.email, lr.firstname or "", lr.lastname or "",
                    lr.phone_mobile or "", lr.phone_home or "", _bool_label(lr.has_parental_authority)]
        else:
            row += ["", "", "", "", "", ""]
    return row


def _send_invitation_email(user, request):
    """Envoie un email d'invitation au nouveau représentant légal avec un lien de définition de mot de passe."""
    form = PasswordResetForm({"email": user.email})
    if form.is_valid():
        form.save(
            request=request,
            use_https=request.is_secure(),
            email_template_name="registration/invitation_email.html",
            subject_template_name="registration/invitation_subject.txt",
        )


def _get_role(user):
    if LegalRepresentative.objects.filter(user=user).exists():
        return "legal_representative"
    if RegistrationSupervisor.objects.filter(user=user).exists():
        return "registration_supervisor"
    return None


class LoginView(APIView):
    permission_classes = (AllowAny,)
    authentication_classes = ()

    def post(self, request):
        username = request.data.get("username", "")
        password = request.data.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is None:
            return Response({"detail": "Identifiants invalides."}, status=status.HTTP_401_UNAUTHORIZED)
        login(request, user)
        get_token(request)  # ensure csrftoken cookie is set on the response
        role = _get_role(user)
        if role is None:
            return Response({"detail": "Aucun rôle attribué à ce compte."}, status=status.HTTP_403_FORBIDDEN)
        return Response({"username": user.username, "role": role})


class LogoutView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def post(self, request):
        logout(request)
        return Response(status=status.HTTP_204_NO_CONTENT)


class MeView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request):
        role = _get_role(request.user)
        return Response({"username": request.user.username, "role": role})


class RegistrationsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    parser_classes = (MultiPartParser, JSONParser)

    def get(self, request):
        if not RegistrationSupervisor.objects.filter(user=request.user).exists():
            return Response({"detail": "Réservé aux gestionnaires."}, status=status.HTTP_403_FORBIDDEN)
        today = date.today()
        campaign = RegistrationCampaign.objects.filter(year__year=today.year).first()
        if not campaign:
            return Response([])
        qs = (
            RegistrationFile.objects
            .filter(campaign=campaign)
            .select_related("pupil")
            .prefetch_related(
                "pupil__pupillegalrepresentative_set__legal_representative"
            )
        )
        serializer = RegistrationListItemSerializer(qs, many=True, context={"request": request})
        return Response(serializer.data)

    def post(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        serializer = RegistrationFileCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        registration_file = serializer.save(legal_representative=lr)
        return Response({"id": registration_file.id}, status=status.HTTP_201_CREATED)


class MyRegistrationsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        pupil_ids = PupilLegalRepresentative.objects.filter(legal_representative=lr).values_list("pupil_id", flat=True)
        qs = (
            RegistrationFile.objects
            .filter(pupil_id__in=pupil_ids)
            .select_related("pupil")
            .prefetch_related("pupil__pupillegalrepresentative_set__legal_representative")
        )
        serializer = RegistrationListItemSerializer(qs, many=True, context={"request": request})
        return Response(serializer.data)


class CoRepresentativeView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        plrs = PupilLegalRepresentative.objects.filter(legal_representative=lr).select_related("pupil")
        result = []
        for plr in plrs:
            pupil = plr.pupil
            co_plr = (
                PupilLegalRepresentative.objects
                .filter(pupil=pupil)
                .exclude(legal_representative=lr)
                .select_related("legal_representative__user")
                .first()
            )
            result.append({
                "pupil_id": pupil.id,
                "pupil_firstname": pupil.firstname,
                "pupil_lastname": pupil.lastname,
                "grade_label": pupil.get_grade_display(),
                "co_representative_email": co_plr.legal_representative.user.email if co_plr else None,
            })
        return Response(result)

    def post(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        pupil_id = request.data.get("pupil_id")
        if not pupil_id:
            return Response({"detail": "pupil_id est requis."}, status=status.HTTP_400_BAD_REQUEST)
        plr = PupilLegalRepresentative.objects.filter(legal_representative=lr, pupil_id=pupil_id).select_related("pupil").first()
        if not plr:
            return Response({"detail": "Élève non associé à votre compte."}, status=status.HTTP_400_BAD_REQUEST)
        pupil = plr.pupil
        if PupilLegalRepresentative.objects.filter(pupil=pupil).exclude(legal_representative=lr).exists():
            return Response({"detail": "Un co-représentant est déjà associé à cet élève."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = LegalRepresentativeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        new_lr = serializer.save()
        PupilLegalRepresentative.objects.create(pupil=pupil, legal_representative=new_lr)
        _send_invitation_email(new_lr.user, request)
        return Response({"id": new_lr.id}, status=status.HTTP_201_CREATED)


class MyProfileView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    parser_classes = (MultiPartParser, JSONParser)

    def get(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        serializer = LegalRepresentativeProfileSerializer(lr, context={"request": request})
        return Response(serializer.data)

    def patch(self, request):
        lr = LegalRepresentative.objects.filter(user=request.user).first()
        if not lr:
            return Response({"detail": "Réservé aux représentants légaux."}, status=status.HTTP_403_FORBIDDEN)
        serializer = LegalRepresentativeProfileSerializer(
            lr, data=request.data, partial=True, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class LegalRepresentativesView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request):
        if not RegistrationSupervisor.objects.filter(user=request.user).exists():
            return Response({"detail": "Réservé aux gestionnaires."}, status=status.HTTP_403_FORBIDDEN)
        qs = LegalRepresentative.objects.select_related("user").order_by("-date_creation")
        serializer = LegalRepresentativeListItemSerializer(qs, many=True)
        return Response(serializer.data)

    def post(self, request):
        if not RegistrationSupervisor.objects.filter(user=request.user).exists():
            return Response({"detail": "Réservé aux gestionnaires."}, status=status.HTTP_403_FORBIDDEN)
        serializer = LegalRepresentativeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lr = serializer.save()
        return Response({"id": lr.id}, status=status.HTTP_201_CREATED)


class RegistrationDetailView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    parser_classes = (MultiPartParser, JSONParser)

    def _get_lr_registration(self, user, pk):
        """Return the RegistrationFile if it belongs to the requesting LR, else None."""
        lr = LegalRepresentative.objects.filter(user=user).first()
        if not lr:
            return None
        pupil_ids = PupilLegalRepresentative.objects.filter(
            legal_representative=lr
        ).values_list("pupil_id", flat=True)
        try:
            return (
                RegistrationFile.objects
                .select_related("pupil")
                .prefetch_related("pupil__pupillegalrepresentative_set__legal_representative__user")
                .get(pk=pk, pupil_id__in=pupil_ids)
            )
        except RegistrationFile.DoesNotExist:
            return None

    def get(self, request, pk):
        if RegistrationSupervisor.objects.filter(user=request.user).exists():
            try:
                registration = (
                    RegistrationFile.objects
                    .select_related("pupil")
                    .prefetch_related("pupil__pupillegalrepresentative_set__legal_representative__user")
                    .get(pk=pk)
                )
            except RegistrationFile.DoesNotExist:
                return Response({"detail": "Dossier introuvable."}, status=status.HTTP_404_NOT_FOUND)
        else:
            registration = self._get_lr_registration(request.user, pk)
            if registration is None:
                return Response({"detail": "Dossier introuvable."}, status=status.HTTP_404_NOT_FOUND)

        serializer = RegistrationDetailSerializer(registration, context={"request": request})
        return Response(serializer.data)

    def patch(self, request, pk):
        # Supervisors can toggle is_closed
        if RegistrationSupervisor.objects.filter(user=request.user).exists():
            try:
                registration = RegistrationFile.objects.get(pk=pk)
            except RegistrationFile.DoesNotExist:
                return Response({"detail": "Dossier introuvable."}, status=status.HTTP_404_NOT_FOUND)
            is_closed = request.data.get("is_closed")
            if is_closed is None:
                return Response({"detail": "is_closed est requis."}, status=status.HTTP_400_BAD_REQUEST)
            registration.is_closed = bool(is_closed)
            registration.save(update_fields=["is_closed"])
            return Response({"id": registration.id, "is_closed": registration.is_closed})

        # Legal representatives can edit their own non-closed registration
        registration = self._get_lr_registration(request.user, pk)
        if registration is None:
            return Response({"detail": "Dossier introuvable."}, status=status.HTTP_404_NOT_FOUND)
        if registration.is_closed:
            return Response(
                {"detail": "Ce dossier est clôturé et ne peut plus être modifié."},
                status=status.HTTP_403_FORBIDDEN,
            )
        serializer = RegistrationFileUpdateSerializer(registration, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"id": registration.id})


class RegistrationsExportView(APIView):
    """Export de tous les dossiers de l'année en cours — superviseurs uniquement."""

    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request):
        if not RegistrationSupervisor.objects.filter(user=request.user).exists():
            return Response({"detail": "Réservé aux gestionnaires."}, status=status.HTTP_403_FORBIDDEN)

        fmt = request.query_params.get("format", "csv")
        if fmt not in ("csv", "excel", "odt", "html"):
            return Response({"detail": "Format invalide. Valeurs acceptées : csv, excel, odt, html."}, status=status.HTTP_400_BAD_REQUEST)

        today = date.today()
        campaign = RegistrationCampaign.objects.filter(year__year=today.year).first()
        registrations = (
            RegistrationFile.objects
            .filter(campaign=campaign)
            .select_related("pupil")
            .prefetch_related("pupil__pupillegalrepresentative_set__legal_representative__user")
            .order_by("pupil__lastname", "pupil__firstname")
        ) if campaign else RegistrationFile.objects.none()

        if fmt == "csv":
            return self._export_csv(registrations)
        if fmt == "excel":
            return self._export_excel(registrations)
        if fmt == "odt":
            return self._export_odt(registrations)
        return self._export_html(registrations, request)

    def _export_csv(self, registrations):
        buf = io.StringIO()
        writer = csv.writer(buf, dialect="excel")
        writer.writerow(EXPORT_HEADERS)
        for reg in registrations:
            writer.writerow(_registration_row(reg))
        content = buf.getvalue().encode("utf-8-sig")  # BOM pour compatibilité Excel
        response = HttpResponse(content, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="inscriptions.csv"'
        return response

    def _export_excel(self, registrations):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inscriptions"
        ws.append(EXPORT_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A56A0", end_color="1A56A0", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True)
        for reg in registrations:
            ws.append(_registration_row(reg))
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="inscriptions.xlsx"'
        return response

    def _export_odt(self, registrations):
        def esc_xml(v):
            return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

        def cell(text, is_header=False):
            style = "Heading" if is_header else "Contents"
            return (
                f'<table:table-cell table:style-name="{style}" office:value-type="string">'
                f'<text:p>{esc_xml(text)}</text:p>'
                f'</table:table-cell>'
            )

        rows_xml = "<table:table-row>" + "".join(cell(h, True) for h in EXPORT_HEADERS) + "</table:table-row>"
        for reg in registrations:
            rows_xml += "<table:table-row>" + "".join(cell(v) for v in _registration_row(reg)) + "</table:table-row>"

        content_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
  xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
  xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"
  xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"
  xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0"
  xmlns:fo="urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"
  office:version="1.2">
<office:automatic-styles>
  <style:style style:name="Heading" style:family="table-cell">
    <style:text-properties fo:font-weight="bold" fo:color="#ffffff"/>
    <style:table-cell-properties fo:background-color="#1a56a0"/>
  </style:style>
  <style:style style:name="Contents" style:family="table-cell"/>
</office:automatic-styles>
<office:body><office:text>
  <table:table table:name="Inscriptions">{rows_xml}</table:table>
</office:text></office:body>
</office:document-content>"""

        manifest_xml = """<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0" manifest:version="1.2">
  <manifest:file-entry manifest:full-path="/" manifest:media-type="application/vnd.oasis.opendocument.text"/>
  <manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>
</manifest:manifest>"""

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("mimetype", "application/vnd.oasis.opendocument.text", compress_type=zipfile.ZIP_STORED)
            zf.writestr("META-INF/manifest.xml", manifest_xml)
            zf.writestr("content.xml", content_xml)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type="application/vnd.oasis.opendocument.text")
        response["Content-Disposition"] = 'attachment; filename="inscriptions.odt"'
        return response

    def _export_html(self, registrations, request):
        from rest_api.html_export import generate_all_registrations_html
        from rest_api.serializers import RegistrationListItemSerializer

        regs_with_completion = [
            (reg, RegistrationListItemSerializer(reg, context={"request": request}).data.get("completion_pct", 0))
            for reg in registrations
        ]
        content = generate_all_registrations_html(regs_with_completion)
        response = HttpResponse(content.encode("utf-8"), content_type="text/html; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="inscriptions.html"'
        return response


def _ascii_filename(name: str) -> str:
    """Normalize a string to ASCII for use in filenames."""
    normalized = unicodedata.normalize("NFKD", name)
    ascii_str = normalized.encode("ascii", "ignore").decode("ascii")
    return "".join(c if (c.isalnum() or c in "-_.") else "_" for c in ascii_str).strip("_")


class RegistrationDownloadView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)

    def get(self, request, pk):
        if not RegistrationSupervisor.objects.filter(user=request.user).exists():
            return Response({"detail": "Réservé aux gestionnaires."}, status=status.HTTP_403_FORBIDDEN)
        try:
            reg = (
                RegistrationFile.objects
                .select_related("pupil")
                .prefetch_related(
                    "pupil__pupillegalrepresentative_set__legal_representative__user"
                )
                .get(pk=pk)
            )
        except RegistrationFile.DoesNotExist:
            return Response({"detail": "Dossier introuvable."}, status=status.HTTP_404_NOT_FOUND)

        # Compute completion for the HTML recap
        from rest_api.serializers import RegistrationListItemSerializer
        tmp_serializer = RegistrationListItemSerializer(reg, context={"request": request})
        completion_pct = tmp_serializer.data.get("completion_pct", 0)

        # Build ZIP in memory
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            # HTML recap
            html_content = generate_registration_html(reg, completion_pct)
            zf.writestr("recap.html", html_content.encode("utf-8"))

            # Registration documents
            doc_slots = [
                ("document",             "certificat_naissance"),
                ("vaccination_document", "carnet_sante"),
                ("insurance_document",   "attestation_assurance"),
                ("divorce_judgment",     "jugement_divorce"),
            ]
            for field_name, base_name in doc_slots:
                file_field = getattr(reg, field_name)
                if not file_field:
                    continue
                ext = file_field.name.rsplit(".", 1)[-1] if "." in file_field.name else "bin"
                zip_name = f"documents/{base_name}.{ext}"
                try:
                    with file_field.open("rb") as f:
                        zf.writestr(zip_name, f.read())
                except Exception:
                    pass

            # Legal representative pool attestations
            plrs = reg.pupil.pupillegalrepresentative_set.all()
            for plr in plrs:
                lr = plr.legal_representative
                if not lr.pool_attestation:
                    continue
                lr_slug = _ascii_filename(
                    f"{lr.firstname or ''}_{lr.lastname or lr.user.email}"
                ).lower()
                ext = lr.pool_attestation.name.rsplit(".", 1)[-1] if "." in lr.pool_attestation.name else "bin"
                zip_name = f"documents/attestation_piscine_{lr_slug}.{ext}"
                try:
                    with lr.pool_attestation.open("rb") as f:
                        zf.writestr(zip_name, f.read())
                except Exception:
                    pass

        buffer.seek(0)
        pupil_slug = _ascii_filename(
            f"{reg.pupil.firstname}_{reg.pupil.lastname}"
        ).lower()
        zip_filename = f"dossier_{pupil_slug}.zip"

        response = HttpResponse(buffer.read(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
        return response